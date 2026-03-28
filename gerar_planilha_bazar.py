"""
Gerador de Planilha Automatizada - Bazar da Thaís
Planilha profissional estilo McKinsey para controle de precificação de maquiagens.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.worksheet.datavalidation import DataValidation
import os

# ── Paleta de cores McKinsey ──
AZUL_ESCURO = "1B2A4A"
AZUL_MEDIO = "2E5090"
AZUL_CLARO = "4472C4"
CINZA_ESCURO = "404040"
CINZA_MEDIO = "808080"
CINZA_CLARO = "F2F2F2"
BRANCO = "FFFFFF"
VERDE = "70AD47"
AMARELO = "FFC000"
LARANJA = "ED7D31"
VERMELHO = "FF4B4B"
ROSA = "E8A5C8"
ROSA_CLARO = "FCE4EC"
DOURADO = "D4AF37"
PRETO = "000000"

# ── Estilos reutilizáveis ──
BORDA_FINA = Border(
    left=Side(style="thin", color=CINZA_MEDIO),
    right=Side(style="thin", color=CINZA_MEDIO),
    top=Side(style="thin", color=CINZA_MEDIO),
    bottom=Side(style="thin", color=CINZA_MEDIO),
)

BORDA_MEDIA = Border(
    left=Side(style="medium", color=AZUL_ESCURO),
    right=Side(style="medium", color=AZUL_ESCURO),
    top=Side(style="medium", color=AZUL_ESCURO),
    bottom=Side(style="medium", color=AZUL_ESCURO),
)

FONT_TITULO = Font(name="Calibri", size=20, bold=True, color=BRANCO)
FONT_SUBTITULO = Font(name="Calibri", size=14, bold=True, color=AZUL_ESCURO)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=BRANCO)
FONT_BODY = Font(name="Calibri", size=11, color=CINZA_ESCURO)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True, color=CINZA_ESCURO)
FONT_MONEY = Font(name="Calibri", size=11, color=CINZA_ESCURO)
FONT_MONEY_BOLD = Font(name="Calibri", size=12, bold=True, color=AZUL_ESCURO)
FONT_SMALL = Font(name="Calibri", size=9, italic=True, color=CINZA_MEDIO)

FILL_HEADER = PatternFill(start_color=AZUL_ESCURO, end_color=AZUL_ESCURO, fill_type="solid")
FILL_HEADER2 = PatternFill(start_color=AZUL_MEDIO, end_color=AZUL_MEDIO, fill_type="solid")
FILL_ZEBRA = PatternFill(start_color=CINZA_CLARO, end_color=CINZA_CLARO, fill_type="solid")
FILL_VERDE = PatternFill(start_color=VERDE, end_color=VERDE, fill_type="solid")
FILL_AMARELO = PatternFill(start_color=AMARELO, end_color=AMARELO, fill_type="solid")
FILL_LARANJA = PatternFill(start_color=LARANJA, end_color=LARANJA, fill_type="solid")
FILL_VERMELHO = PatternFill(start_color=VERMELHO, end_color=VERMELHO, fill_type="solid")
FILL_ROSA = PatternFill(start_color=ROSA_CLARO, end_color=ROSA_CLARO, fill_type="solid")
FILL_TITULO = PatternFill(start_color=AZUL_ESCURO, end_color=AZUL_ESCURO, fill_type="solid")
FILL_DOURADO = PatternFill(start_color=DOURADO, end_color=DOURADO, fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

BRL_FORMAT = 'R$ #,##0.00'
PCT_FORMAT = '0%'
DATE_FORMAT = 'DD/MM/YYYY'


def apply_style(cell, font=None, fill=None, alignment=None, border=None, number_format=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    if number_format: cell.number_format = number_format


def style_range(ws, row, col_start, col_end, **kwargs):
    for c in range(col_start, col_end + 1):
        apply_style(ws.cell(row=row, column=c), **kwargs)


def create_title_bar(ws, row, col_start, col_end, title, subtitle=None):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=title)
    apply_style(cell, font=FONT_TITULO, fill=FILL_TITULO, alignment=ALIGN_CENTER)
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).fill = FILL_TITULO
        ws.cell(row=row, column=c).border = BORDA_MEDIA
    if subtitle:
        ws.merge_cells(start_row=row+1, start_column=col_start, end_row=row+1, end_column=col_end)
        sub = ws.cell(row=row+1, column=col_start, value=subtitle)
        apply_style(sub, font=FONT_SMALL, fill=FILL_ROSA, alignment=ALIGN_CENTER)
        for c in range(col_start, col_end + 1):
            ws.cell(row=row+1, column=c).fill = FILL_ROSA


def create_header_row(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=col_start + i, value=h)
        apply_style(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER, border=BORDA_FINA)


def create_section_header(ws, row, col_start, col_end, text):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=text)
    apply_style(cell, font=Font(name="Calibri", size=12, bold=True, color=BRANCO),
                fill=FILL_HEADER2, alignment=ALIGN_CENTER)
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).fill = FILL_HEADER2
        ws.cell(row=row, column=c).border = BORDA_FINA


# ═══════════════════════════════════════════
# ABA 1: CATÁLOGO (Principal)
# ═══════════════════════════════════════════
def create_catalogo(wb):
    ws = wb.active
    ws.title = "Catálogo"
    ws.sheet_properties.tabColor = AZUL_ESCURO

    # Larguras das colunas
    widths = {
        'A': 5, 'B': 18, 'C': 30, 'D': 15, 'E': 12,
        'F': 10, 'G': 15, 'H': 15, 'I': 15,
        'J': 15, 'K': 15, 'L': 15, 'M': 15,
        'N': 15, 'O': 15, 'P': 15, 'Q': 18,
        'R': 15, 'S': 30
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Título
    create_title_bar(ws, 1, 1, 19, "🌸 BAZAR DA THAÍS — CATÁLOGO DE MAQUIAGENS",
                     "Planilha automatizada de precificação • Consultoria McKinsey Style")

    # Headers
    headers = [
        "#", "Marca", "Produto", "Cor/Tom", "Tipo",
        "Volume", "Categoria Risco", "Preço Mín\nMercado", "Preço Máx\nMercado",
        "Preço Médio\nMercado", "MEDIANA\n(Referência)", "Fontes\nPesquisadas", "Data\nPesquisa",
        "🟢 Nunca\nUsado", "🟡 25%\nUsado", "🟠 50%\nUsado", "🔴 75%\nUsado",
        "Ajuste\nCategoria", "Observações"
    ]
    create_header_row(ws, 3, headers)
    ws.row_dimensions[3].height = 45

    # Data validation para Tipo
    tipo_val = DataValidation(
        type="list",
        formula1='"Base,Batom,Blush,Bronzer,Contorno,Corretivo,Delineador,Gloss,Iluminador,Lápis,Máscara Cílios,Paleta Sombras,Pó Compacto,Pó Solto,Primer,Protetor Labial,Sombra,Spray Fixador,Outro"',
        allow_blank=True
    )
    tipo_val.error = "Selecione um tipo válido"
    tipo_val.errorTitle = "Tipo Inválido"
    ws.add_data_validation(tipo_val)

    # Data validation para Categoria de Risco
    risco_val = DataValidation(
        type="list",
        formula1='"Normal,Alto Risco Sanitário,Lacrado/Selado,Edição Limitada,Vencimento Próximo,Drugstore (<R$50)"',
        allow_blank=True
    )
    ws.add_data_validation(risco_val)

    # Preencher 50 linhas com fórmulas
    for row in range(4, 54):
        idx = row - 3
        ws.cell(row=row, column=1, value=idx)
        apply_style(ws.cell(row=row, column=1), font=FONT_BODY, alignment=ALIGN_CENTER, border=BORDA_FINA)

        # Colunas de input (B-M)
        for c in range(2, 14):
            cell = ws.cell(row=row, column=c)
            apply_style(cell, font=FONT_BODY, alignment=ALIGN_LEFT if c <= 6 else ALIGN_CENTER, border=BORDA_FINA)
            if c in (8, 9, 10, 11):  # Preços de mercado
                cell.number_format = BRL_FORMAT
            if c == 13:  # Data
                cell.number_format = DATE_FORMAT

        # Validações
        tipo_val.add(ws.cell(row=row, column=5))
        risco_val.add(ws.cell(row=row, column=7))

        r = str(row)

        # Coluna N: Preço Nunca Usado (com ajuste de categoria)
        # Fórmula: Se mediana preenchida → Mediana × 0.75 × (1 + Ajuste)
        formula_nunca = (
            f'=IF(K{r}="","",ROUND(K{r}*0.75*(1+R{r}),2))'
        )
        cell_n = ws.cell(row=row, column=14, value=formula_nunca)
        apply_style(cell_n, font=FONT_MONEY_BOLD, alignment=ALIGN_CENTER, border=BORDA_FINA,
                    number_format=BRL_FORMAT)
        cell_n.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

        # Coluna O: 25% Usado
        formula_25 = f'=IF(N{r}="","",ROUND(N{r}*0.85,2))'
        cell_o = ws.cell(row=row, column=15, value=formula_25)
        apply_style(cell_o, font=FONT_MONEY, alignment=ALIGN_CENTER, border=BORDA_FINA,
                    number_format=BRL_FORMAT)
        cell_o.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

        # Coluna P: 50% Usado
        formula_50 = f'=IF(N{r}="","",ROUND(N{r}*0.70,2))'
        cell_p = ws.cell(row=row, column=16, value=formula_50)
        apply_style(cell_p, font=FONT_MONEY, alignment=ALIGN_CENTER, border=BORDA_FINA,
                    number_format=BRL_FORMAT)
        cell_p.fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")

        # Coluna Q: 75% Usado
        formula_75 = f'=IF(N{r}="","",ROUND(N{r}*0.50,2))'
        cell_q = ws.cell(row=row, column=17, value=formula_75)
        apply_style(cell_q, font=FONT_MONEY, alignment=ALIGN_CENTER, border=BORDA_FINA,
                    number_format=BRL_FORMAT)
        cell_q.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

        # Coluna R: Ajuste de Categoria (fórmula baseada na coluna G)
        formula_ajuste = (
            f'=IF(G{r}="Alto Risco Sanitário",-0.15,'
            f'IF(G{r}="Vencimento Próximo",-0.20,'
            f'IF(G{r}="Lacrado/Selado",0.05,'
            f'IF(G{r}="Edição Limitada",0.07,'
            f'IF(G{r}="Drugstore (<R$50)",0.13,0)))))'
        )
        cell_r = ws.cell(row=row, column=18, value=formula_ajuste)
        apply_style(cell_r, font=FONT_BODY, alignment=ALIGN_CENTER, border=BORDA_FINA,
                    number_format=PCT_FORMAT)

        # Coluna S: Observações
        apply_style(ws.cell(row=row, column=19), font=FONT_SMALL, alignment=ALIGN_LEFT, border=BORDA_FINA)

        # Zebra striping
        if idx % 2 == 0:
            for c in [1, 2, 3, 4, 5, 6, 7, 12, 13, 18, 19]:
                ws.cell(row=row, column=c).fill = FILL_ZEBRA

    # Congelar painéis
    ws.freeze_panes = "B4"

    # Filtro automático
    ws.auto_filter.ref = "A3:S53"

    return ws


# ═══════════════════════════════════════════
# ABA 2: PESQUISA DE MERCADO
# ═══════════════════════════════════════════
def create_pesquisa(wb):
    ws = wb.create_sheet("Pesquisa de Mercado")
    ws.sheet_properties.tabColor = AZUL_MEDIO

    widths = {'A': 5, 'B': 25, 'C': 25, 'D': 20, 'E': 15, 'F': 35, 'G': 15}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    create_title_bar(ws, 1, 1, 7, "🔍 PESQUISA DE MERCADO — DETALHAMENTO",
                     "Registro detalhado de preços por loja para cada produto")

    # Template para 10 produtos (cada um com espaço para 14 lojas)
    current_row = 4
    for prod in range(1, 11):
        # Seção do produto
        create_section_header(ws, current_row, 1, 7, f"PRODUTO #{prod}")
        current_row += 1

        # Referência ao catálogo
        ref_cell = ws.cell(row=current_row, column=1, value="Produto:")
        apply_style(ref_cell, font=FONT_BODY_BOLD, alignment=ALIGN_RIGHT)
        ref_val = ws.cell(row=current_row, column=2, value=f"=Catálogo!C{prod+3}")
        apply_style(ref_val, font=FONT_BODY_BOLD, alignment=ALIGN_LEFT)

        marca_cell = ws.cell(row=current_row, column=3, value="Marca:")
        apply_style(marca_cell, font=FONT_BODY_BOLD, alignment=ALIGN_RIGHT)
        marca_val = ws.cell(row=current_row, column=4, value=f"=Catálogo!B{prod+3}")
        apply_style(marca_val, font=FONT_BODY_BOLD, alignment=ALIGN_LEFT)
        current_row += 1

        # Headers da tabela de lojas
        headers = ["#", "Loja", "Produto Encontrado", "Tipo Loja", "Preço (R$)", "Link/Observação", "Data"]
        create_header_row(ws, current_row, headers)
        current_row += 1

        # 14 linhas para lojas
        lojas_padrao = [
            "Sephora Brasil", "Beleza na Web", "Época Cosméticos", "Loja Oficial da Marca",
            "O Boticário", "Quem Disse Berenice", "Natura", "Drogasil",
            "Droga Raia", "Panvel", "Renner Beauty", "C&A Beauty",
            "Amazon Brasil", "Mercado Livre (oficial)"
        ]
        for i, loja in enumerate(lojas_padrao):
            ws.cell(row=current_row, column=1, value=i+1)
            apply_style(ws.cell(row=current_row, column=1), font=FONT_BODY, alignment=ALIGN_CENTER, border=BORDA_FINA)

            ws.cell(row=current_row, column=2, value=loja)
            for c in range(2, 8):
                cell = ws.cell(row=current_row, column=c)
                apply_style(cell, font=FONT_BODY, alignment=ALIGN_LEFT if c != 5 else ALIGN_CENTER, border=BORDA_FINA)
                if c == 5:
                    cell.number_format = BRL_FORMAT
                if c == 7:
                    cell.number_format = DATE_FORMAT

            if i % 2 == 0:
                for c in range(1, 8):
                    ws.cell(row=current_row, column=c).fill = FILL_ZEBRA
            current_row += 1

        # Linha de resumo
        first_price_row = current_row - 14
        last_price_row = current_row - 1
        rng = f"E{first_price_row}:E{last_price_row}"

        resumo = [
            ("Menor:", f"=IF(COUNTA({rng})=0,\"\",MIN({rng}))"),
            ("Maior:", f"=IF(COUNTA({rng})=0,\"\",MAX({rng}))"),
            ("Média:", f"=IF(COUNTA({rng})=0,\"\",AVERAGE({rng}))"),
            ("MEDIANA:", f"=IF(COUNTA({rng})=0,\"\",MEDIAN({rng}))"),
        ]
        for label, formula in resumo:
            ws.cell(row=current_row, column=4, value=label)
            apply_style(ws.cell(row=current_row, column=4), font=FONT_BODY_BOLD, alignment=ALIGN_RIGHT)
            cell = ws.cell(row=current_row, column=5, value=formula)
            is_mediana = "MEDIANA" in label
            apply_style(cell,
                        font=FONT_MONEY_BOLD if is_mediana else FONT_MONEY,
                        alignment=ALIGN_CENTER,
                        border=BORDA_MEDIA if is_mediana else BORDA_FINA,
                        number_format=BRL_FORMAT)
            if is_mediana:
                cell.fill = PatternFill(start_color="DCEDC8", end_color="DCEDC8", fill_type="solid")
            current_row += 1

        current_row += 2  # Espaço entre produtos

    ws.freeze_panes = "A4"
    return ws


# ═══════════════════════════════════════════
# ABA 3: DASHBOARD
# ═══════════════════════════════════════════
def create_dashboard(wb):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = DOURADO

    widths = {'A': 3, 'B': 25, 'C': 18, 'D': 18, 'E': 18, 'F': 18, 'G': 18, 'H': 18, 'I': 18, 'J': 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    create_title_bar(ws, 1, 1, 10, "📊 DASHBOARD — VISÃO EXECUTIVA DO BAZAR",
                     "Resumo automático • Atualiza conforme produtos são adicionados")

    # ── KPIs ──
    row = 4
    create_section_header(ws, row, 2, 10, "INDICADORES-CHAVE")
    row += 1

    kpis = [
        ("Total de Produtos", '=COUNTA(Catálogo!B4:B53)', None),
        ("Valor Estoque (Mercado)", '=SUMPRODUCT((Catálogo!K4:K53<>"")*Catálogo!K4:K53)', BRL_FORMAT),
        ("Valor Bazar (Nunca Usado)", '=SUMPRODUCT((Catálogo!N4:N53<>"")*Catálogo!N4:N53)', BRL_FORMAT),
        ("Valor Bazar (25% Usado)", '=SUMPRODUCT((Catálogo!O4:O53<>"")*Catálogo!O4:O53)', BRL_FORMAT),
        ("Valor Bazar (50% Usado)", '=SUMPRODUCT((Catálogo!P4:P53<>"")*Catálogo!P4:P53)', BRL_FORMAT),
        ("Valor Bazar (75% Usado)", '=SUMPRODUCT((Catálogo!Q4:Q53<>"")*Catálogo!Q4:Q53)', BRL_FORMAT),
        ("Economia Média p/ Comprador", '=IF(C6=0,"",1-C7/C6)', PCT_FORMAT),
        ("Ticket Médio (Nunca Usado)", '=IF(C5=0,"",C7/C5)', BRL_FORMAT),
    ]

    for i, (label, formula, fmt) in enumerate(kpis):
        r = row + i
        ws.cell(row=r, column=2, value=label)
        apply_style(ws.cell(row=r, column=2), font=FONT_BODY_BOLD, alignment=ALIGN_LEFT, border=BORDA_FINA)
        cell = ws.cell(row=r, column=3, value=formula)
        apply_style(cell, font=FONT_MONEY_BOLD, alignment=ALIGN_CENTER, border=BORDA_FINA)
        if fmt:
            cell.number_format = fmt

        # Fundo alternado
        if i % 2 == 0:
            ws.cell(row=r, column=2).fill = FILL_ZEBRA
            ws.cell(row=r, column=3).fill = FILL_ZEBRA

    row += len(kpis) + 2

    # ── Distribuição por Tipo ──
    create_section_header(ws, row, 2, 5, "DISTRIBUIÇÃO POR TIPO DE PRODUTO")
    row += 1

    tipos = [
        "Base", "Batom", "Blush", "Bronzer", "Contorno", "Corretivo",
        "Delineador", "Gloss", "Iluminador", "Lápis", "Máscara Cílios",
        "Paleta Sombras", "Pó Compacto", "Pó Solto", "Primer", "Sombra",
        "Spray Fixador", "Outro"
    ]

    ws.cell(row=row, column=2, value="Tipo")
    ws.cell(row=row, column=3, value="Qtd")
    ws.cell(row=row, column=4, value="Valor Mercado")
    ws.cell(row=row, column=5, value="Valor Bazar")
    for c in range(2, 6):
        apply_style(ws.cell(row=row, column=c), font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER, border=BORDA_FINA)
    row += 1

    tipos_start_row = row
    for i, tipo in enumerate(tipos):
        r = row + i
        ws.cell(row=r, column=2, value=tipo)
        apply_style(ws.cell(row=r, column=2), font=FONT_BODY, alignment=ALIGN_LEFT, border=BORDA_FINA)

        ws.cell(row=r, column=3, value=f'=COUNTIF(Catálogo!E4:E53,B{r})')
        apply_style(ws.cell(row=r, column=3), font=FONT_BODY, alignment=ALIGN_CENTER, border=BORDA_FINA)

        ws.cell(row=r, column=4, value=f'=SUMPRODUCT((Catálogo!E4:E53=B{r})*Catálogo!K4:K53)')
        apply_style(ws.cell(row=r, column=4), font=FONT_MONEY, alignment=ALIGN_CENTER, border=BORDA_FINA,
                    number_format=BRL_FORMAT)

        ws.cell(row=r, column=5, value=f'=SUMPRODUCT((Catálogo!E4:E53=B{r})*Catálogo!N4:N53)')
        apply_style(ws.cell(row=r, column=5), font=FONT_MONEY, alignment=ALIGN_CENTER, border=BORDA_FINA,
                    number_format=BRL_FORMAT)

        if i % 2 == 0:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = FILL_ZEBRA

    tipos_end_row = row + len(tipos) - 1
    row = tipos_end_row + 2

    # ── Distribuição por Marca ──
    create_section_header(ws, row, 2, 5, "TOP MARCAS (por valor de mercado)")
    row += 1
    ws.cell(row=row, column=2, value="Marca")
    ws.cell(row=row, column=3, value="Qtd")
    ws.cell(row=row, column=4, value="Valor Mercado")
    ws.cell(row=row, column=5, value="Valor Bazar")
    for c in range(2, 6):
        apply_style(ws.cell(row=row, column=c), font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER, border=BORDA_FINA)
    row += 1
    # 15 linhas para marcas (preenchidas manualmente ou pelo agente)
    for i in range(15):
        r = row + i
        for c in range(2, 6):
            apply_style(ws.cell(row=r, column=c), font=FONT_BODY, alignment=ALIGN_CENTER if c >= 3 else ALIGN_LEFT, border=BORDA_FINA)
            if c == 4 or c == 5:
                ws.cell(row=r, column=c).number_format = BRL_FORMAT
        if i % 2 == 0:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = FILL_ZEBRA

    row += 17

    # ── Tabela de Fórmulas (referência) ──
    create_section_header(ws, row, 2, 6, "REFERÊNCIA DE FÓRMULAS")
    row += 1
    formulas_ref = [
        ("Preço de Referência", "Mediana dos preços de mercado", "=MEDIAN(preços)"),
        ("🟢 Bazar Nunca Usado", "25% abaixo do mercado + ajuste", "Referência × 0,75 × (1 + Ajuste)"),
        ("🟡 Bazar 25% Usado", "15% abaixo do preço nunca usado", "Nunca Usado × 0,85"),
        ("🟠 Bazar 50% Usado", "30% abaixo do preço nunca usado", "Nunca Usado × 0,70"),
        ("🔴 Bazar 75% Usado", "50% abaixo do preço nunca usado", "Nunca Usado × 0,50"),
        ("Ajuste: Alto Risco", "Máscara, delineador líq., gloss", "-15%"),
        ("Ajuste: Vencimento Próximo", "< 6 meses de validade", "-20%"),
        ("Ajuste: Lacrado/Selado", "Produto nunca aberto", "+5%"),
        ("Ajuste: Edição Limitada", "Descontinuado com demanda", "+7%"),
        ("Ajuste: Drugstore", "Preço original < R$50", "+13% (desconto menor)"),
    ]
    ws.cell(row=row, column=2, value="Item")
    ws.cell(row=row, column=3, value="Descrição")
    ws.cell(row=row, column=4, value="Fórmula")
    for c in range(2, 5):
        apply_style(ws.cell(row=row, column=c), font=FONT_HEADER, fill=FILL_HEADER2, alignment=ALIGN_CENTER, border=BORDA_FINA)
    row += 1
    for i, (item, desc, form) in enumerate(formulas_ref):
        r = row + i
        ws.cell(row=r, column=2, value=item)
        ws.cell(row=r, column=3, value=desc)
        ws.cell(row=r, column=4, value=form)
        for c in range(2, 5):
            apply_style(ws.cell(row=r, column=c), font=FONT_BODY, alignment=ALIGN_LEFT, border=BORDA_FINA)
        if i % 2 == 0:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = FILL_ZEBRA

    ws.freeze_panes = "A4"
    return ws


# ═══════════════════════════════════════════
# ABA 4: ETIQUETAS DE PREÇO
# ═══════════════════════════════════════════
def create_etiquetas(wb):
    ws = wb.create_sheet("Etiquetas")
    ws.sheet_properties.tabColor = ROSA

    ws.column_dimensions['A'].width = 3
    for col in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 22

    create_title_bar(ws, 1, 1, 5, "🏷️ ETIQUETAS DE PREÇO PARA IMPRESSÃO",
                     "Recorte e cole nos produtos • Atualiza automaticamente do catálogo")

    row = 4
    for prod in range(1, 26):
        cat_row = prod + 3

        # Borda da etiqueta
        for r in range(row, row + 6):
            for c in range(2, 6):
                ws.cell(row=r, column=c).border = BORDA_MEDIA

        # Linha 1: Nome do produto
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        cell = ws.cell(row=row, column=2, value=f'=IF(Catálogo!B{cat_row}="","",Catálogo!B{cat_row}&" — "&Catálogo!C{cat_row})')
        apply_style(cell, font=Font(name="Calibri", size=13, bold=True, color=AZUL_ESCURO),
                    fill=FILL_ROSA, alignment=ALIGN_CENTER)
        for c in range(2, 6):
            ws.cell(row=row, column=c).fill = FILL_ROSA

        # Linha 2: Cor e tipo
        ws.merge_cells(start_row=row+1, start_column=2, end_row=row+1, end_column=5)
        ws.cell(row=row+1, column=2, value=f'=IF(Catálogo!D{cat_row}="","",Catálogo!D{cat_row}&" | "&Catálogo!E{cat_row}&" | "&Catálogo!F{cat_row})')
        apply_style(ws.cell(row=row+1, column=2), font=FONT_SMALL, alignment=ALIGN_CENTER)

        # Linha 3: Preço de mercado (riscado)
        ws.merge_cells(start_row=row+2, start_column=2, end_row=row+2, end_column=5)
        ws.cell(row=row+2, column=2, value=f'=IF(Catálogo!K{cat_row}="","","De: "&TEXT(Catálogo!K{cat_row},"R$ #.##0,00"))')
        apply_style(ws.cell(row=row+2, column=2),
                    font=Font(name="Calibri", size=10, strikethrough=True, color=CINZA_MEDIO),
                    alignment=ALIGN_CENTER)

        # Linha 4: Preços do bazar
        labels = ["🟢 Novo", "🟡 25%", "🟠 50%", "🔴 75%"]
        cols_ref = [14, 15, 16, 17]  # N, O, P, Q
        fills = [
            PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
            PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
            PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
            PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
        ]
        for i, (label, col_idx, fill) in enumerate(zip(labels, cols_ref, fills)):
            c = 2 + i
            col_letter = get_column_letter(col_idx)
            ws.cell(row=row+3, column=c, value=label)
            apply_style(ws.cell(row=row+3, column=c),
                        font=Font(name="Calibri", size=8, bold=True, color=CINZA_ESCURO),
                        alignment=ALIGN_CENTER, fill=fill)

            ws.cell(row=row+4, column=c, value=f'=IF(Catálogo!{col_letter}{cat_row}="","",Catálogo!{col_letter}{cat_row})')
            apply_style(ws.cell(row=row+4, column=c),
                        font=Font(name="Calibri", size=14, bold=True, color=AZUL_ESCURO),
                        alignment=ALIGN_CENTER, fill=fill,
                        number_format=BRL_FORMAT)

        # Linha 5: @bazar da thais
        ws.merge_cells(start_row=row+5, start_column=2, end_row=row+5, end_column=5)
        ws.cell(row=row+5, column=2, value="🌸 Bazar da Thaís")
        apply_style(ws.cell(row=row+5, column=2),
                    font=Font(name="Calibri", size=9, italic=True, color=ROSA),
                    alignment=ALIGN_CENTER,
                    fill=PatternFill(start_color=AZUL_ESCURO, end_color=AZUL_ESCURO, fill_type="solid"))
        for c in range(2, 6):
            ws.cell(row=row+5, column=c).fill = PatternFill(start_color=AZUL_ESCURO, end_color=AZUL_ESCURO, fill_type="solid")

        row += 8  # Espaço entre etiquetas

    return ws


# ═══════════════════════════════════════════
# ABA 5: GUIA DE USO
# ═══════════════════════════════════════════
def create_guia(wb):
    ws = wb.create_sheet("Como Usar")
    ws.sheet_properties.tabColor = VERDE

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 80

    create_title_bar(ws, 1, 1, 2, "📖 GUIA DE USO — BAZAR DA THAÍS", "")

    instrucoes = [
        ("🚀 COMO FUNCIONA", [
            "1. Abra o VS Code com esta pasta do projeto",
            "2. No chat do Copilot, selecione o agente 'Bazar da Thaís' (@ → Bazar da Thaís)",
            "3. Envie a foto da maquiagem no chat",
            "4. O agente identifica o produto, pesquisa preços e sugere valores",
            "5. Peça para o agente preencher a planilha automaticamente",
        ]),
        ("📋 ABAS DA PLANILHA", [
            "• Catálogo: Tabela principal com todos os produtos e preços calculados automaticamente",
            "• Pesquisa de Mercado: Detalhamento de preços por loja para cada produto",
            "• Dashboard: Visão executiva com KPIs, totais e distribuição",
            "• Etiquetas: Etiquetas de preço prontas para imprimir e colar nos produtos",
            "• Como Usar: Este guia",
        ]),
        ("💰 COMO OS PREÇOS SÃO CALCULADOS", [
            "1. Pesquisamos o preço em 5-14 lojas confiáveis (Sephora, Beleza na Web, etc.)",
            "2. A MEDIANA dos preços vira o Preço de Referência",
            "3. Preço de Bazar (Nunca Usado) = Referência × 75% (sempre mais barato que loja)",
            "4. 25% Usado = Nunca Usado × 85%",
            "5. 50% Usado = Nunca Usado × 70%",
            "6. 75% Usado = Nunca Usado × 50%",
            "7. Ajustes automáticos são aplicados por categoria de risco",
        ]),
        ("⚠️ AJUSTES DE CATEGORIA", [
            "• Normal: sem ajuste",
            "• Alto Risco Sanitário (máscara, delineador, gloss): -15%",
            "• Vencimento Próximo (< 6 meses): -20%",
            "• Lacrado/Selado: +5%",
            "• Edição Limitada: +7%",
            "• Drugstore (< R$50): +13% (desconto menor do mercado)",
        ]),
        ("🎯 DICAS DE VENDA", [
            "• Use preços psicológicos: R$ 29,90 em vez de R$ 30,00",
            "• Monte kits: 'Leve 3 pague 2'",
            "• Crie urgência: 'Últimas unidades' / 'Edição que saiu de linha'",
            "• Poste as etiquetas no Instagram/WhatsApp para divulgar",
            "• Aceite Pix e facilite o pagamento",
        ]),
    ]

    row = 4
    for titulo, items in instrucoes:
        create_section_header(ws, row, 1, 2, titulo)
        row += 1
        for item in items:
            ws.cell(row=row, column=2, value=item)
            apply_style(ws.cell(row=row, column=2), font=FONT_BODY, alignment=ALIGN_LEFT, border=BORDA_FINA)
            row += 1
        row += 1

    return ws


# ═══════════════════════════════════════════
# GERADOR PRINCIPAL
# ═══════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()

    create_catalogo(wb)
    create_pesquisa(wb)
    create_dashboard(wb)
    create_etiquetas(wb)
    create_guia(wb)

    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Bazar_Thais_Maquiagem.xlsx")
    wb.save(output_path)
    print(f"✅ Planilha gerada com sucesso: {output_path}")
    print(f"   📊 5 abas: Catálogo | Pesquisa de Mercado | Dashboard | Etiquetas | Como Usar")
    print(f"   📋 50 linhas de produtos prontas com fórmulas automáticas")
    print(f"   🏷️ 25 etiquetas de preço auto-preenchíveis")


if __name__ == "__main__":
    main()
