import openpyxl
wb = openpyxl.load_workbook(r"c:\Users\guskramer\Personal\Bazar_Thais_Maquiagem.xlsx")
ws = wb["Catálogo"]

# Clear display headers
headers = ["Produto", "Menor", "Mediana", "Ajuste", "Nunca", "25%", "50%", "75%"]
print(f"{'Produto':<25} | {'Menor':<8} | {'Mediana':<8} | {'Ajuste':<7} | {'Nunca':<7} | {'25%':<7} | {'50%':<7} | {'75%':<7}")
print("-" * 85)

# Produto 1: Tarte (Row 4)
prod1 = ws["C4"].value[:20] if ws["C4"].value else "N/A"
menor1 = ws["H4"].value if ws["H4"].value else 0
mediana1 = ws["K4"].value if ws["K4"].value else 0
ajuste1 = ws["R4"].value if ws["R4"].value else 0
nunca1 = ws["N4"].value if ws["N4"].value else "calc..."
uso25_1 = ws["O4"].value if ws["O4"].value else "calc..."
uso50_1 = ws["P4"].value if ws["P4"].value else "calc..."
uso75_1 = ws["Q4"].value if ws["Q4"].value else "calc..."

print(f"{prod1:<25} | {menor1:<8.0f} | {mediana1:<8.0f} | {ajuste1*100:<6.1f}% | {nunca1:<7.0f} | {uso25_1:<7.0f} | {uso50_1:<7.0f} | {uso75_1:<7.0f}")

# Produto 2: Eucerin (Row 5)
prod2 = ws["C5"].value[:20] if ws["C5"].value else "N/A"
menor2 = ws["H5"].value if ws["H5"].value else 0
mediana2 = ws["K5"].value if ws["K5"].value else 0
ajuste2 = ws["R5"].value if ws["R5"].value else 0
nunca2 = ws["N5"].value if ws["N5"].value else "calc..."
uso25_2 = ws["O5"].value if ws["O5"].value else "calc..."
uso50_2 = ws["P5"].value if ws["P5"].value else "calc..."
uso75_2 = ws["Q5"].value if ws["Q5"].value else "calc..."

print(f"{prod2:<25} | {menor2:<8.0f} | {mediana2:<8.0f} | {ajuste2*100:<6.1f}% | {nunca2:<7.0f} | {uso25_2:<7.0f} | {uso50_2:<7.0f} | {uso75_2:<7.0f}")

wb.close()
